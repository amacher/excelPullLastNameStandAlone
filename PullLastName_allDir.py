from openpyxl import Workbook
from openpyxl import load_workbook
import os
from tkinter import Tk
from tkinter import BOTH
from tkinter import Label
from tkinter import filedialog 
# Go on from there; f is a handle to the file that the user picked

root = Tk()
# gui = Tk()
root.title('Excel file-Last Name Column')
#The info to include in the window
Label(
    root,
    text="Select file in folder...\nThis window will close when done.",
    font=('Times',20)
    ).pack(fill=BOTH, expand=True)
#Asks for the file to start the processing
root.filename = filedialog.askopenfilename(initialdir="", title="Select A File", filetypes=[("xlsx files", "*.xlsx"), ("all files", "*.*")])

#The extra info at end to be removed
SUFFIX = ['JR', 'JR.', 'SR', 'SR.', 'LTD', 'II', 'PC']
#to have all citys in the city col put here.
CITY = []

#Pulls the path for where the files are located.
path = os.path.split(root.filename)[0]
#The extensions you need to check for to run.
ext = ('.xlsx')

# Function to take everything to right of last ' ' 
# and put that into LastName everything to left goes in FirstName
def splitName(fullName):
    name = fullName.rsplit(' ', 1)
    first = name[0]
    last = name[-1]
    #save first/last to the new columns (H, I)
    ws.cell(row = r, column = 8).value = str(last)
    ws.cell(row = r, column = 9).value = str(first)
    wb.save(FILE_NAME)

#For loop to go through all files in the directory
for file in os.listdir(path):
    #checking if the extension is in file name if not it will go to next step and not be processed.
    if file.endswith(ext):
        #assigning new file name for saving of file
        newFile = file.rsplit('.', 1)
        FILE_NAME = path+'/'+newFile[0]+'_update'+'.'+newFile[1]
        #Open Excel file, change later to take what is dropped onto it.
        wb = load_workbook(path+'/'+file)
        #if only one worksheet this will open just that, another ws needed this will need to be changed.
        ws = wb.active
        #adds a new column header in the H1 spot
        ws['H1'] = 'Last Name'
        ws['I1'] = 'First Name'
        #PULL all the towns from the CITY column (D), don't pull doubles
        column = ws['D']
        #putting into a set to automatically only keeping one of each; doing before going into loop since this only needs to run once per WB.
        CITY = set([column[x].value for x in range(len(column))])
        #Iterate through each row
        for r in range(2, ws.max_row + 1):
            fullName = ws.cell(row = r, column = 1).value
            company = ws.cell(row = r, column = 2).value
            
        #IF Company != '' move that to new Last Name col
            if company != '': #or company is not None:
                ws.cell(row = r, column = 8).value = company
                wb.save(FILE_NAME)
            # Else from Full Name col (A)
            else:
                # IF contains only 1 ' ' then process it (function to split first and last)
                if fullName.count(' ') == 1:
                    splitName(fullName)
                #Check for more than 1 ' '
                else:
                    #Doesn't work for towns with a space 'St. Louis'
                    querywords = fullName.rsplit(' ')
                    #pulls last 2 words in list
                    townCheck = ' '.join(querywords[-2:])
                    if townCheck.upper() in CITY:
                        #removes the last 2 words if it's in Town (checks for 2 word towns)
                        querywords = querywords[: len(querywords) -2]
                    # print(querywords)
                    #Check if the last word after the last ' ' matches any from CITY, if so then delete
                    resultwords  = [word for word in querywords if word.upper() not in CITY]
                    #Check if the last word is SUFFIX, if so delete
                    resultwords2  = [word for word in resultwords if word.upper() not in SUFFIX]
                    #Joins the list back together
                    result2 = ' '.join(resultwords2)
                    splitName(result2)
    else:
        continue

##Commenting out so when the file is done processing the app will close.        
# root.mainloop()